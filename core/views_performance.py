"""Views for the enhanced collection performance dashboard.

This module provides an improved implementation of the collection
performance dashboard and its supporting APIs.  In addition to the
stacked bar chart previously available, the new dashboard offers a
donut chart illustrating interview contributions per interviewer for
the selected project, a daily trend line chart and a sortable table
highlighting the top interviewers.  An extended Excel export is also
provided, including a raw data sheet listing every call.

The views defined here mirror the names of the previous endpoints so
that existing URLs continue to work.  If you wish to customise the
behaviour further, adjust the aggregation logic or chart data as
needed.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Dict, List, Sequence

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Avg, Count, DurationField, ExpressionWrapper, F, Q
from django.db.models.functions import TruncDate, ExtractHour
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone

from .models import Interview, Project
# Reuse helper functions from the main views module.  In addition to
# _user_has_panel and _user_is_organisation we also import
# _get_accessible_projects so we can filter interview data to only
# projects the current user is permitted to view.
from .views import (
    _user_has_panel,
    _user_is_organisation,
    _get_accessible_projects,
    _get_locked_projects,
    _localise_text,
)


def _parse_datetime_param(value: str | None) -> datetime | None:
    """Return a timezone-aware datetime parsed from an ISO string."""

    if not value:
        return None
    try:
        dt = datetime.fromisoformat(value)
    except ValueError:
        return None
    if timezone.is_naive(dt):
        return timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


def _normalise_multi_value(raw_values: Sequence[str]) -> List[str]:
    """Expand comma-separated values in list-style query parameters."""

    expanded: List[str] = []
    for value in raw_values:
        if not value:
            continue
        if ',' in value:
            expanded.extend([part.strip() for part in value.split(',') if part.strip()])
        else:
            expanded.append(value.strip())
    return expanded


def _extract_filters(
    request: HttpRequest,
    accessible_projects: List[Project],
) -> Dict[str, object]:
    """Parse query parameters shared by the dashboard and export views."""

    accessible_ids = {proj.id for proj in accessible_projects}

    raw_project_params = request.GET.getlist('projects')
    if not raw_project_params and request.GET.get('project'):
        raw_project_params = [request.GET['project']]
    project_ids: List[int] = []
    invalid_projects: List[int] = []
    for value in _normalise_multi_value(raw_project_params):
        try:
            project_id = int(value)
        except (TypeError, ValueError):
            continue
        if project_id in accessible_ids and project_id not in project_ids:
            project_ids.append(project_id)
        else:
            invalid_projects.append(project_id)
    if invalid_projects:
        raise PermissionError('project_locked')

    raw_user_params = request.GET.getlist('users')
    if not raw_user_params and request.GET.get('users'):
        raw_user_params = [request.GET['users']]
    user_ids: List[int] = []
    for value in _normalise_multi_value(raw_user_params):
        try:
            user_id = int(value)
        except (TypeError, ValueError):
            continue
        if user_id not in user_ids:
            user_ids.append(user_id)

    start_raw = request.GET.get('start_date')
    end_raw = request.GET.get('end_date')

    return {
        'project_ids': project_ids,
        'user_ids': user_ids,
        'start_dt': _parse_datetime_param(start_raw),
        'end_dt': _parse_datetime_param(end_raw),
        'start_raw': start_raw or '',
        'end_raw': end_raw or '',
    }


def _filter_interviews(
    user: User,
    accessible_projects: List[Project],
    filters: Dict[str, object],
):
    """Return an Interview queryset filtered by the supplied parameters."""

    project_ids: List[int] = filters.get('project_ids', [])  # type: ignore[assignment]
    user_ids: List[int] = filters.get('user_ids', [])  # type: ignore[assignment]
    start_dt: datetime | None = filters.get('start_dt')  # type: ignore[assignment]
    end_dt: datetime | None = filters.get('end_dt')  # type: ignore[assignment]

    qs = Interview.objects.select_related('project', 'user').filter(project__in=accessible_projects)
    if project_ids:
        qs = qs.filter(project__id__in=project_ids)
    if start_dt:
        qs = qs.filter(created_at__gte=start_dt)
    if end_dt:
        qs = qs.filter(created_at__lte=end_dt)
    if not _user_is_organisation(user):
        qs = qs.filter(user=user)
    elif user_ids:
        qs = qs.filter(user__id__in=user_ids)
    return qs


def _build_chart_payload(qs) -> Dict[str, object]:
    """Aggregate interview data into chart/table structures."""

    bar_labels: List[str] = []
    bar_totals: List[int] = []
    bar_successes: List[int] = []
    bar_projects: List[Dict[str, object]] = []
    for row in (
        qs.values('project__id', 'project__name')
        .annotate(total=Count('id'), success=Count('id', filter=Q(status=True)))
        .order_by('project__name')
    ):
        label = row['project__name'] or ''
        bar_labels.append(label)
        bar_totals.append(row['total'])
        bar_successes.append(row['success'])
        bar_projects.append({
            'id': row['project__id'],
            'name': label,
            'total': row['total'],
            'success': row['success'],
        })

    donut_labels: List[str] = []
    donut_values: List[int] = []
    donut_segments: List[Dict[str, object]] = []
    for row in (
        qs.values('project__name', 'user__id', 'user__first_name')
        .annotate(total=Count('id'))
        .order_by('-total', 'project__name', 'user__first_name')
    ):
        user_label = row['user__first_name'] or str(row['user__id'])
        project_label = row['project__name'] or ''
        combined_label = f"{project_label} — {user_label}" if project_label else user_label
        donut_labels.append(combined_label)
        donut_values.append(row['total'])
        donut_segments.append({
            'project': project_label,
            'user': user_label,
            'label': combined_label,
            'value': row['total'],
        })

    daily_labels: List[str] = []
    daily_totals: List[int] = []
    daily_successes: List[int] = []
    for row in (
        qs.annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(total=Count('id'), success=Count('id', filter=Q(status=True)))
        .order_by('day')
    ):
        day = row['day']
        day_str = day.isoformat() if hasattr(day, 'isoformat') else str(day)
        daily_labels.append(day_str)
        daily_totals.append(row['total'])
        daily_successes.append(row['success'])

    top_rows: List[Dict[str, object]] = []
    for row in (
        qs.values('project__id', 'project__name', 'user__id', 'user__first_name')
        .annotate(total=Count('id'), success=Count('id', filter=Q(status=True)))
        .order_by('-total', 'project__name', 'user__first_name')
    ):
        total = row['total']
        success = row['success']
        rate = (success / total * 100) if total else 0
        top_rows.append({
            'project_id': row['project__id'],
            'project': row['project__name'] or '',
            'user_id': row['user__id'],
            'user': row['user__first_name'] or str(row['user__id']),
            'total_calls': total,
            'successful_calls': success,
            'success_rate': round(rate, 2),
        })

    status_rows = qs.values('status').annotate(count=Count('id'))
    status_breakdown = {
        'successful': 0,
        'unsuccessful': 0,
        'unknown': 0,
    }
    for row in status_rows:
        status_value = row['status']
        if status_value is True:
            status_breakdown['successful'] = row['count']
        elif status_value is False:
            status_breakdown['unsuccessful'] = row['count']
        else:
            status_breakdown['unknown'] = row['count']

    code_labels: List[str] = []
    code_values: List[int] = []
    code_items: List[Dict[str, object]] = []
    for row in qs.values('code').annotate(count=Count('id')).order_by('code'):
        code_value = row['code']
        label = str(code_value) if code_value is not None else '—'
        code_labels.append(label)
        code_values.append(row['count'])
        code_items.append({
            'code': code_value,
            'label': label,
            'count': row['count'],
        })

    duration_qs = qs.filter(start_form__isnull=False, end_form__isnull=False)
    duration_expr = ExpressionWrapper(F('end_form') - F('start_form'), output_field=DurationField())
    avg_duration = duration_qs.aggregate(avg_duration=Avg(duration_expr))['avg_duration']
    avg_duration_minutes: float | None = None
    avg_duration_label = ''
    if avg_duration:
        total_seconds = avg_duration.total_seconds()
        avg_duration_minutes = total_seconds / 60 if total_seconds else 0
        seconds = int(total_seconds % 60)
        minutes = int((total_seconds // 60) % 60)
        hours = int(total_seconds // 3600)
        if hours:
            avg_duration_label = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        else:
            avg_duration_label = f"{minutes:02d}:{seconds:02d}"

    hourly_rows = (
        qs.annotate(hour=ExtractHour('created_at'))
        .values('hour')
        .annotate(total=Count('id'), success=Count('id', filter=Q(status=True)))
    )
    hour_map = {row['hour']: row for row in hourly_rows if row['hour'] is not None}
    hourly_labels: List[str] = []
    hourly_totals: List[int] = []
    hourly_successes: List[int] = []
    for hour in range(24):
        hourly_labels.append(f"{hour:02d}:00")
        row = hour_map.get(hour)
        if row:
            hourly_totals.append(row['total'])
            hourly_successes.append(row['success'])
        else:
            hourly_totals.append(0)
            hourly_successes.append(0)

    total_interviews = sum(bar_totals)
    successful_interviews = sum(bar_successes)
    success_rate = (successful_interviews / total_interviews * 100) if total_interviews else 0

    peak_hour = None
    peak_hour_label = ''
    if any(hourly_totals):
        max_total = max(hourly_totals)
        if max_total > 0:
            peak_index = hourly_totals.index(max_total)
            peak_hour = peak_index
            peak_hour_label = f"{peak_index:02d}:00"

    return {
        'bar': {
            'labels': bar_labels,
            'totals': bar_totals,
            'successes': bar_successes,
            'projects': bar_projects,
        },
        'donut': {
            'labels': donut_labels,
            'values': donut_values,
            'segments': donut_segments,
        },
        'daily': {
            'labels': daily_labels,
            'totals': daily_totals,
            'successes': daily_successes,
        },
        'top': {
            'rows': top_rows,
        },
        'codes': {
            'labels': code_labels,
            'values': code_values,
            'items': code_items,
        },
        'hourly': {
            'labels': hourly_labels,
            'totals': hourly_totals,
            'successes': hourly_successes,
        },
        'meta': {
            'total_interviews': total_interviews,
            'successful_interviews': successful_interviews,
            'success_rate': round(success_rate, 2),
            'status_breakdown': status_breakdown,
            'code_breakdown': code_items,
            'average_duration_minutes': avg_duration_minutes,
            'average_duration_label': avg_duration_label,
            'duration_sample_size': duration_qs.count(),
            'peak_hour': peak_hour,
            'peak_hour_label': peak_hour_label,
        },
    }

# Attempt to import openpyxl for Excel export
try:
    import openpyxl  # type: ignore
    from openpyxl.chart import BarChart, Reference, PieChart
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color='1f2937', end_color='1f2937', fill_type='solid')
    KPI_FILL = PatternFill(start_color='0f172a', end_color='0f172a', fill_type='solid')
    KPI_FONT = Font(color='ffffff', bold=True)
    STRIPE_FILLS = [
        PatternFill(start_color='0b1220', end_color='0b1220', fill_type='solid'),
        PatternFill(start_color='111827', end_color='111827', fill_type='solid'),
    ]
    BORDER_STYLE = Border(
        left=Side(style='thin', color='1f2937'),
        right=Side(style='thin', color='1f2937'),
        top=Side(style='thin', color='1f2937'),
        bottom=Side(style='thin', color='1f2937'),
    )
    DATA_ALIGNMENT = Alignment(vertical='center', wrap_text=False)
except Exception:
    openpyxl = None  # type: ignore
    BarChart = Reference = PieChart = None  # type: ignore
    Alignment = Border = Font = PatternFill = Side = None  # type: ignore
    get_column_letter = None  # type: ignore
    HEADER_FILL = KPI_FILL = KPI_FONT = BORDER_STYLE = DATA_ALIGNMENT = None  # type: ignore
    STRIPE_FILLS = []  # type: ignore


def _style_tabular_section(
    worksheet,
    header_row: int,
    data_start: int,
    data_end: int,
    start_col: int = 1,
    end_col: int | None = None,
) -> None:
    """Apply header formatting, zebra rows, and borders to a table section."""

    if end_col is None:
        end_col = worksheet.max_column
    if data_end < data_start:
        return
    for col_idx in range(start_col, end_col + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        if HEADER_FILL:
            cell.fill = HEADER_FILL
        if BORDER_STYLE:
            cell.border = BORDER_STYLE
        cell.font = Font(bold=True, color='ffffff') if Font else None
        if DATA_ALIGNMENT:
            cell.alignment = DATA_ALIGNMENT
    stripe_count = len(STRIPE_FILLS)
    for row_idx in range(data_start, data_end + 1):
        fill = None
        if stripe_count:
            fill = STRIPE_FILLS[(row_idx - data_start) % stripe_count]
        for col_idx in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill
            if BORDER_STYLE:
                cell.border = BORDER_STYLE
            if DATA_ALIGNMENT:
                cell.alignment = DATA_ALIGNMENT


def _auto_fit_columns(worksheet) -> None:
    """Best-effort auto-fit for Excel columns."""

    if get_column_letter is None:
        return
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            if len(text) > max_length:
                max_length = len(text)
        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[column_letter].width = max(12, adjusted_width)

@login_required
def collection_performance(request: HttpRequest) -> HttpResponse:
    """Render the enhanced collection performance dashboard.

    This view prepares the list of projects and users accessible to the
    current user and renders the dashboard template.  Only users with
    the ``collection_performance`` panel permission may access this
    page.  Organisation accounts see all projects they participate in,
    whereas individual interviewers only see their own projects.
    """
    user = request.user
    if not _user_has_panel(user, 'collection_performance'):
        messages.error(request, 'Access denied: you do not have collection performance permissions.')
        return redirect('home')
    # Determine accessible projects: organisations see all their
    # membership projects; individuals see only their own.
    # Build the list of projects for which the user has the collection_performance panel permission.
    accessible_projects = _get_accessible_projects(user, panel='collection_performance')
    if not accessible_projects:
        locked = _get_locked_projects(user, panel='collection_performance')
        if locked:
            locked_names = ', '.join(sorted({p.name for p in locked}))
            messages.error(
                request,
                f'Access denied: project deadlines have passed ({locked_names}). Only the owner can continue to view the collection performance dashboard.',
            )
        else:
            messages.error(request, 'Access denied: there are no projects available for collection performance analytics.')
        return redirect('home')
    try:
        filters = _extract_filters(request, accessible_projects)
    except PermissionError:
        messages.error(request, 'Requested project is not available for this dashboard.')
        return redirect('collection_performance')

    if _user_is_organisation(user):
        users_qs = User.objects.filter(memberships__project__in=accessible_projects).distinct()
        selected_user_ids = filters['user_ids']  # type: ignore[index]
    else:
        users_qs = User.objects.filter(pk=user.pk)
        selected_user_ids = [user.pk]

    context = {
        'projects': sorted(accessible_projects, key=lambda p: p.name.lower()),
        'users': users_qs.order_by('first_name'),
        'filters': {
            'project_ids': filters['project_ids'],  # type: ignore[index]
            'user_ids': selected_user_ids,
            'start': filters['start_raw'],  # type: ignore[index]
            'end': filters['end_raw'],  # type: ignore[index]
        },
    }
    return render(request, 'collection_performance.html', context)


@login_required
def collection_performance_data(request: HttpRequest) -> JsonResponse:
    """Return aggregated interview statistics for performance charts."""

    user = request.user
    if not _user_has_panel(user, 'collection_performance'):
        return JsonResponse({'error': 'forbidden'}, status=403)

    accessible_projects = _get_accessible_projects(user, panel='collection_performance')
    if not accessible_projects:
        return JsonResponse(
            {
                'bar': {'labels': [], 'totals': [], 'successes': [], 'projects': []},
                'donut': {'labels': [], 'values': [], 'segments': []},
                'daily': {'labels': [], 'totals': [], 'successes': []},
                'top': {'rows': []},
                'codes': {'labels': [], 'values': [], 'items': []},
                'hourly': {'labels': [], 'totals': [], 'successes': []},
                'meta': {
                    'total_interviews': 0,
                    'successful_interviews': 0,
                    'success_rate': 0,
                    'status_breakdown': {'successful': 0, 'unsuccessful': 0, 'unknown': 0},
                    'code_breakdown': [],
                    'average_duration_minutes': None,
                    'average_duration_label': '',
                    'duration_sample_size': 0,
                    'peak_hour': None,
                    'peak_hour_label': '',
                },
            }
        )

    try:
        filters = _extract_filters(request, accessible_projects)
    except PermissionError:
        return JsonResponse({'error': 'project_locked'}, status=403)

    qs = _filter_interviews(user, accessible_projects, filters)
    payload = _build_chart_payload(qs)
    payload['filters'] = {
        'projects': filters['project_ids'],  # type: ignore[index]
        'users': filters['user_ids'],  # type: ignore[index]
        'start_date': filters['start_raw'],  # type: ignore[index]
        'end_date': filters['end_raw'],  # type: ignore[index]
    }
    return JsonResponse(payload)


@login_required
def collection_performance_raw(request: HttpRequest) -> JsonResponse:
    """Return paginated raw interview records for the dashboard table."""

    user = request.user
    if not _user_has_panel(user, 'collection_performance'):
        return JsonResponse({'error': 'forbidden'}, status=403)

    accessible_projects = _get_accessible_projects(user, panel='collection_performance')
    if not accessible_projects:
        return JsonResponse(
            {
                'results': [],
                'page': 1,
                'page_size': 0,
                'total_pages': 0,
                'total_items': 0,
            }
        )

    try:
        filters = _extract_filters(request, accessible_projects)
    except PermissionError:
        return JsonResponse({'error': 'project_locked'}, status=403)

    qs = (
        _filter_interviews(user, accessible_projects, filters)
        .select_related('project', 'user')
        .order_by('-created_at')
    )

    allowed_page_sizes = {30, 50, 200}
    default_page_size = 30
    try:
        requested_size = int(request.GET.get('page_size', default_page_size))
    except (TypeError, ValueError):
        requested_size = default_page_size
    page_size = requested_size if requested_size in allowed_page_sizes else default_page_size

    paginator = Paginator(qs, page_size)
    if paginator.count == 0:
        return JsonResponse(
            {
                'results': [],
                'page': 1,
                'page_size': page_size,
                'total_pages': 1,
                'total_items': 0,
            }
        )
    try:
        page_number = int(request.GET.get('page', 1))
    except (TypeError, ValueError):
        page_number = 1
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_number = 1
        page_obj = paginator.page(page_number)
    except EmptyPage:
        page_number = paginator.num_pages or 1
        page_obj = paginator.page(page_number)

    def _format_dt(value):
        if not value:
            return None
        return timezone.localtime(value).isoformat()

    results = [
        {
            'id': interview.id,
            'project_id': interview.project_id,
            'project': interview.project.name if interview.project else '',
            'user_id': interview.user_id,
            'user': interview.user.first_name or interview.user.get_username(),
            'code': interview.code,
            'status': interview.status,
            'start_form': _format_dt(interview.start_form),
            'end_form': _format_dt(interview.end_form),
            'created_at': _format_dt(interview.created_at),
        }
        for interview in page_obj
    ]

    return JsonResponse(
        {
            'results': results,
            'page': page_obj.number,
            'page_size': page_size,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
        }
    )

@login_required
def collection_performance_export(request: HttpRequest) -> HttpResponse:
    """Export the filtered collection performance dataset to Excel."""

    user = request.user
    lang = request.session.get('lang', 'en')
    if not _user_has_panel(user, 'collection_performance'):
        message = _localise_text(
            lang,
            'Access denied: you do not have collection performance permissions.',
            'دسترسی مجاز نیست: شما به این داشبورد دسترسی ندارید.',
        )
        messages.error(request, message)
        return redirect('home')
    if openpyxl is None:
        message = _localise_text(
            lang,
            'Excel export is not available on this server.',
            'امکان تهیه فایل اکسل روی این سرور وجود ندارد.',
        )
        return JsonResponse({'error': message}, status=501)

    accessible_projects = _get_accessible_projects(user, panel='collection_performance')
    if not accessible_projects:
        message = _localise_text(
            lang,
            'There are no projects available for collection performance analytics.',
            'پروژه‌ای برای گزارش عملکرد جمع‌آوری در دسترس نیست.',
        )
        messages.error(request, message)
        return redirect('collection_performance')

    try:
        filters = _extract_filters(request, accessible_projects)
    except PermissionError:
        message = _localise_text(
            lang,
            'Requested project is not available for export.',
            'پروژه انتخاب‌شده برای خروجی مجاز نیست.',
        )
        messages.error(request, message)
        return redirect('collection_performance')

    qs = _filter_interviews(user, accessible_projects, filters)
    payload = _build_chart_payload(qs)
    qs_rows = (
        qs.select_related('project', 'user', 'person')
        .prefetch_related('person__mobiles')
        .order_by('created_at')
    )

    project_names = [
        project.name
        for project in accessible_projects
        if project.id in filters['project_ids']  # type: ignore[index]
    ]
    if not project_names:
        project_names = [project.name for project in accessible_projects]

    user_ids: List[int] = filters['user_ids']  # type: ignore[assignment]
    user_names = list(
        User.objects.filter(id__in=user_ids).values_list('first_name', flat=True)
    ) if user_ids else []

    start_raw = filters['start_raw']  # type: ignore[index]
    end_raw = filters['end_raw']  # type: ignore[index]
    if start_raw and end_raw:
        date_label = f"{start_raw} → {end_raw}"
    elif start_raw or end_raw:
        date_label = start_raw or end_raw or ''
    else:
        date_label = _localise_text(lang, 'All dates', 'همه تاریخ‌ها')

    def _sheet_title(label: str) -> str:
        return label[:31]

    cover_title = _localise_text(lang, 'Cover', 'راهنما')
    summary_title = _localise_text(lang, 'Summary', 'خلاصه')
    top_title = _localise_text(lang, 'Top Interviewers', 'برترین مصاحبه‌گران')
    codes_title = _localise_text(lang, 'Code Breakdown', 'تفکیک کدها')
    raw_title = _localise_text(lang, 'Raw Calls', 'داده خام')

    workbook = openpyxl.Workbook()
    cover_sheet = workbook.active
    cover_sheet.title = _sheet_title(cover_title)

    heading_text = _localise_text(
        lang,
        'Collection Performance Export',
        'خروجی عملکرد جمع‌آوری',
    )
    cover_sheet['A1'] = heading_text
    cover_sheet['A1'].font = Font(bold=True, size=16) if Font else None
    cover_note = _localise_text(
        lang,
        'This workbook summarises the dashboard KPIs, charts, and raw interviews that match your current filters.',
        'این فایل اکسل شاخص‌ها، نمودارها و داده خام مطابق فیلترهای فعال شما را ارائه می‌کند.',
    )
    cover_sheet['A2'] = cover_note
    if DATA_ALIGNMENT:
        cover_sheet['A2'].alignment = Alignment(wrap_text=True)

    generated_label = _localise_text(lang, 'Generated at', 'زمان ایجاد')
    filter_label = _localise_text(lang, 'Filters', 'فیلترها')
    projects_label = _localise_text(lang, 'Projects', 'پروژه‌ها')
    users_label = _localise_text(lang, 'Interviewers', 'مصاحبه‌گران')
    dates_label = _localise_text(lang, 'Date range', 'بازه زمانی')
    all_users_text = _localise_text(lang, 'All interviewers', 'همه مصاحبه‌گران')
    filter_table_start = 4
    cover_sheet.cell(row=filter_table_start, column=1, value=generated_label)
    cover_sheet.cell(
        row=filter_table_start,
        column=2,
        value=timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M'),
    )
    filter_header_row = filter_table_start + 1
    cover_sheet.append([filter_label, _localise_text(lang, 'Value', 'مقدار')])
    cover_sheet.append([projects_label, ', '.join(project_names)])
    cover_sheet.append([users_label, ', '.join(user_names) if user_names else all_users_text])
    cover_sheet.append([dates_label, date_label])
    _style_tabular_section(
        cover_sheet,
        header_row=filter_header_row,
        data_start=filter_header_row + 1,
        data_end=filter_header_row + 3,
    )

    sheet_label = _localise_text(lang, 'Sheet', 'برگه')
    desc_label = _localise_text(lang, 'Description', 'توضیحات')
    summary_desc = _localise_text(
        lang,
        'KPIs, status/code tables, and per-project chart.',
        'شاخص‌ها، جدول وضعیت/کد و نمودار پروژه‌ای.',
    )
    top_desc = _localise_text(
        lang,
        'Ranked interviewer totals with chart.',
        'رتبه‌بندی مصاحبه‌گران همراه با نمودار.',
    )
    codes_desc = _localise_text(
        lang,
        'Code distribution table and chart.',
        'تفکیک کدها با جدول و نمودار.',
    )
    raw_desc = _localise_text(
        lang,
        'Filtered raw interviews with phone and timestamp metadata.',
        'داده خام فیلترشده شامل شماره تماس و زمان ثبت.',
    )
    cover_sheet.append([])
    sheet_table_header = cover_sheet.max_row + 1
    cover_sheet.append([sheet_label, desc_label])
    cover_sheet.append([summary_title, summary_desc])
    cover_sheet.append([top_title, top_desc])
    cover_sheet.append([codes_title, codes_desc])
    cover_sheet.append([raw_title, raw_desc])
    _style_tabular_section(
        cover_sheet,
        header_row=sheet_table_header,
        data_start=sheet_table_header + 1,
        data_end=cover_sheet.max_row,
    )
    cover_sheet.freeze_panes = 'A2'
    _auto_fit_columns(cover_sheet)

    summary_sheet = workbook.create_sheet(title=_sheet_title(summary_title))
    summary_heading = _localise_text(lang, 'Performance summary', 'جمع‌بندی عملکرد')
    summary_sheet['A1'] = summary_heading
    summary_sheet['A1'].font = Font(bold=True, size=14) if Font else None
    filters_text = _localise_text(
        lang,
        'Projects: {projects} | Interviewers: {users} | Dates: {dates}',
        'پروژه‌ها: {projects} | مصاحبه‌گران: {users} | بازه: {dates}',
    ).format(
        projects=', '.join(project_names),
        users=', '.join(user_names) if user_names else all_users_text,
        dates=date_label,
    )
    summary_sheet['A2'] = filters_text
    if DATA_ALIGNMENT:
        summary_sheet['A2'].alignment = Alignment(wrap_text=True)

    summary_sheet.append([])
    kpi_header_row = summary_sheet.max_row + 1
    summary_sheet.append([
        _localise_text(lang, 'Total calls', 'کل تماس‌ها'),
        _localise_text(lang, 'Successful calls', 'تماس‌های موفق'),
        _localise_text(lang, 'Success rate', 'نرخ موفقیت'),
        _localise_text(lang, 'Peak hour', 'ساعت اوج'),
        _localise_text(lang, 'Avg. duration', 'میانگین مدت‌زمان'),
    ])
    kpi_values_row = summary_sheet.max_row + 1
    meta = payload['meta']
    summary_sheet.append([
        meta.get('total_interviews', 0),
        meta.get('successful_interviews', 0),
        f"{meta.get('success_rate', 0)}%",
        meta.get('peak_hour_label') or _localise_text(lang, 'N/A', 'نامشخص'),
        meta.get('average_duration_label') or _localise_text(lang, 'N/A', 'نامشخص'),
    ])
    for col in range(1, 6):
        header_cell = summary_sheet.cell(row=kpi_header_row, column=col)
        value_cell = summary_sheet.cell(row=kpi_values_row, column=col)
        if KPI_FILL:
            value_cell.fill = KPI_FILL
        if KPI_FONT:
            value_cell.font = KPI_FONT
        if HEADER_FILL:
            header_cell.fill = HEADER_FILL
            header_cell.font = Font(bold=True, color='ffffff') if Font else None
        if DATA_ALIGNMENT:
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            value_cell.alignment = Alignment(horizontal='center', vertical='center')

    summary_sheet.append([])
    status_header_row = summary_sheet.max_row + 1
    summary_sheet.append([
        _localise_text(lang, 'Status', 'وضعیت'),
        _localise_text(lang, 'Count', 'تعداد'),
    ])
    status_map = [
        (_localise_text(lang, 'Successful', 'موفق'), 'successful'),
        (_localise_text(lang, 'Unsuccessful', 'ناموفق'), 'unsuccessful'),
        (_localise_text(lang, 'Unknown', 'نامشخص'), 'unknown'),
    ]
    for label, key in status_map:
        summary_sheet.append([label, meta.get('status_breakdown', {}).get(key, 0)])
    _style_tabular_section(
        summary_sheet,
        header_row=status_header_row,
        data_start=status_header_row + 1,
        data_end=summary_sheet.max_row,
        end_col=2,
    )

    summary_sheet.append([])
    code_header_row = summary_sheet.max_row + 1
    summary_sheet.append([
        _localise_text(lang, 'Code', 'کد'),
        _localise_text(lang, 'Label', 'برچسب'),
        _localise_text(lang, 'Count', 'تعداد'),
        _localise_text(lang, 'Share (%)', 'سهم (%)'),
    ])
    code_items = meta.get('code_breakdown') or []
    total_codes = sum(item.get('count', 0) for item in code_items) or 1
    for item in code_items:
        share = round((item.get('count', 0) / total_codes) * 100, 2)
        summary_sheet.append([
            item.get('code'),
            item.get('label'),
            item.get('count'),
            share,
        ])
    _style_tabular_section(
        summary_sheet,
        header_row=code_header_row,
        data_start=code_header_row + 1,
        data_end=summary_sheet.max_row,
        end_col=4,
    )

    summary_sheet.append([])
    project_header_row = summary_sheet.max_row + 1
    summary_sheet.append([
        _localise_text(lang, 'Project', 'پروژه'),
        _localise_text(lang, 'Total calls', 'کل تماس‌ها'),
        _localise_text(lang, 'Successful calls', 'تماس‌های موفق'),
    ])
    for project in payload['bar']['projects']:
        summary_sheet.append([project['name'], project['total'], project['success']])
    project_data_end = summary_sheet.max_row
    _style_tabular_section(
        summary_sheet,
        header_row=project_header_row,
        data_start=project_header_row + 1,
        data_end=project_data_end,
        end_col=3,
    )
    if project_data_end > project_header_row:
        chart = BarChart()
        chart.title = _localise_text(lang, 'Interview performance per project', 'عملکرد پروژه‌ای تماس‌ها')
        chart.x_axis.title = _localise_text(lang, 'Project', 'پروژه')
        chart.y_axis.title = _localise_text(lang, 'Calls', 'تعداد تماس')
        data_ref = Reference(
            summary_sheet,
            min_col=2,
            min_row=project_header_row,
            max_col=3,
            max_row=project_data_end,
        )
        cat_ref = Reference(
            summary_sheet,
            min_col=1,
            min_row=project_header_row + 1,
            max_row=project_data_end,
        )
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.width = 20
        chart.height = 10
        summary_sheet.add_chart(chart, f'E{kpi_header_row}')
    summary_sheet.freeze_panes = 'A2'
    _auto_fit_columns(summary_sheet)

    top_sheet = workbook.create_sheet(title=_sheet_title(top_title))
    top_sheet.append([
        _localise_text(lang, 'Project', 'پروژه'),
        _localise_text(lang, 'Interviewer', 'مصاحبه‌گر'),
        _localise_text(lang, 'Total calls', 'کل تماس‌ها'),
        _localise_text(lang, 'Successful calls', 'تماس‌های موفق'),
        _localise_text(lang, 'Success rate (%)', 'نرخ موفقیت (%)'),
    ])
    for row in payload['top']['rows']:
        top_sheet.append([
            row['project'],
            row['user'],
            row['total_calls'],
            row['successful_calls'],
            row['success_rate'],
        ])
    _style_tabular_section(
        top_sheet,
        header_row=1,
        data_start=2,
        data_end=top_sheet.max_row,
        end_col=5,
    )
    if top_sheet.max_row > 1:
        chart = BarChart()
        chart.title = _localise_text(lang, 'Top interviewers', 'برترین مصاحبه‌گران')
        chart.x_axis.title = _localise_text(lang, 'Interviewer', 'مصاحبه‌گر')
        chart.y_axis.title = _localise_text(lang, 'Calls', 'تعداد تماس')
        data_ref = Reference(top_sheet, min_col=3, min_row=1, max_col=4, max_row=top_sheet.max_row)
        cat_ref = Reference(top_sheet, min_col=2, min_row=2, max_row=top_sheet.max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.width = 18
        chart.height = 9
        top_sheet.add_chart(chart, 'G2')
    top_sheet.freeze_panes = 'A2'
    top_sheet.auto_filter.ref = top_sheet.dimensions
    _auto_fit_columns(top_sheet)

    codes_sheet = workbook.create_sheet(title=_sheet_title(codes_title))
    codes_sheet.append([
        _localise_text(lang, 'Code', 'کد'),
        _localise_text(lang, 'Label', 'برچسب'),
        _localise_text(lang, 'Count', 'تعداد'),
        _localise_text(lang, 'Share (%)', 'سهم (%)'),
    ])
    for item in code_items:
        share = round((item.get('count', 0) / total_codes) * 100, 2)
        codes_sheet.append([
            item.get('code'),
            item.get('label'),
            item.get('count'),
            share,
        ])
    _style_tabular_section(
        codes_sheet,
        header_row=1,
        data_start=2,
        data_end=codes_sheet.max_row,
        end_col=4,
    )
    if codes_sheet.max_row > 1:
        pie = PieChart()
        pie.title = _localise_text(lang, 'Code distribution', 'توزیع کدها')
        labels_ref = Reference(codes_sheet, min_col=2, min_row=2, max_row=codes_sheet.max_row)
        data_ref = Reference(codes_sheet, min_col=3, min_row=1, max_row=codes_sheet.max_row)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels_ref)
        pie.width = 14
        pie.height = 8
        codes_sheet.add_chart(pie, 'F2')
    codes_sheet.freeze_panes = 'A2'
    codes_sheet.auto_filter.ref = codes_sheet.dimensions
    _auto_fit_columns(codes_sheet)

    raw_sheet = workbook.create_sheet(title=_sheet_title(raw_title))
    raw_sheet.append([
        _localise_text(lang, 'Logged at', 'زمان ثبت'),
        _localise_text(lang, 'Project', 'پروژه'),
        _localise_text(lang, 'Interviewer', 'مصاحبه‌گر'),
        _localise_text(lang, 'Phone', 'شماره تماس'),
        _localise_text(lang, 'Code', 'کد'),
        _localise_text(lang, 'Status', 'وضعیت'),
        _localise_text(lang, 'City', 'شهر'),
        _localise_text(lang, 'Age', 'سن'),
        _localise_text(lang, 'Birth year', 'سال تولد'),
        _localise_text(lang, 'Gender', 'جنسیت'),
        _localise_text(lang, 'Form started', 'شروع فرم'),
        _localise_text(lang, 'Form submitted', 'پایان فرم'),
    ])
    success_text = _localise_text(lang, 'Successful', 'موفق')
    failure_text = _localise_text(lang, 'Unsuccessful', 'ناموفق')
    unknown_text = _localise_text(lang, 'Unknown', 'نامشخص')
    for interview in qs_rows:
        phone = ''
        if interview.person and hasattr(interview.person, 'mobiles'):
            mobile = interview.person.mobiles.first()
            if mobile:
                phone = mobile.mobile
        if interview.status is True:
            status_value = success_text
        elif interview.status is False:
            status_value = failure_text
        else:
            status_value = unknown_text
        start_form_str = interview.start_form.isoformat(sep=' ') if interview.start_form else ''
        end_form_str = interview.end_form.isoformat(sep=' ') if interview.end_form else ''
        raw_sheet.append([
            interview.created_at.isoformat(sep=' '),
            interview.project.name if interview.project else '',
            interview.user.first_name or interview.user.get_username(),
            phone,
            interview.code,
            status_value,
            interview.city or '',
            interview.age if interview.age is not None else '',
            interview.birth_year if interview.birth_year is not None else '',
            _localise_text(lang, 'Female', 'زن') if interview.gender is True else (
                _localise_text(lang, 'Male', 'مرد') if interview.gender is False else ''
            ),
            start_form_str,
            end_form_str,
        ])
    _style_tabular_section(
        raw_sheet,
        header_row=1,
        data_start=2,
        data_end=raw_sheet.max_row,
        end_col=12,
    )
    raw_sheet.freeze_panes = 'A2'
    raw_sheet.auto_filter.ref = raw_sheet.dimensions
    _auto_fit_columns(raw_sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = 'collection_performance.xlsx'
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

