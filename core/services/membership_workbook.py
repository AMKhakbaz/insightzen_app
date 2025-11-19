"""Helpers for exporting and importing membership Excel workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Iterable, List, Mapping, Sequence, Tuple

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import transaction

from core.models import Membership, Project

try:  # pragma: no cover - optional dependency
    from openpyxl import Workbook, load_workbook  # type: ignore
except Exception:  # pragma: no cover - handled at runtime
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore

PANEL_FIELDS: Sequence[str] = [
    'database_management',
    'quota_management',
    'collection_management',
    'collection_performance',
    'telephone_interviewer',
    'fieldwork_interviewer',
    'focus_group_panel',
    'qc_management',
    'qc_performance',
    'voice_review',
    'callback_qc',
    'coding',
    'statistical_health_check',
    'tabulation',
    'statistics',
    'funnel_analysis',
    'conjoint_analysis',
    'segmentation_analysis',
]

COLUMN_DEFINITIONS: Sequence[Tuple[str, str]] = [
    ('email', 'Email'),
    ('project', 'Project'),
    ('title', 'Membership Title'),
    ('is_owner', 'Is Owner'),
    *[(field, field.replace('_', ' ').title()) for field in PANEL_FIELDS],
]

BOOLEAN_COLUMNS = {'is_owner', *PANEL_FIELDS}

TRUE_VALUES = {'1', 'true', 'yes', 'y', 'on', 't', '✓', '✔'}
FALSE_VALUES = {'0', 'false', 'no', 'n', 'off', 'f', ''}


class MembershipWorkbookError(Exception):
    """Raised when workbook import/export fails."""


@dataclass
class WorkbookImportResult:
    created: int = 0
    replaced: int = 0
    errors: List[str] = None  # type: ignore[assignment]

    def __post_init__(self) -> None:
        if self.errors is None:
            self.errors = []


def _require_openpyxl() -> None:
    if Workbook is None or load_workbook is None:
        raise MembershipWorkbookError('openpyxl is required to work with membership Excel files.')


def export_memberships_workbook(memberships: Iterable[Membership]) -> BytesIO:
    """Return a BytesIO containing the workbook for the provided memberships."""

    _require_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Memberships'
    ws.append([label for _, label in COLUMN_DEFINITIONS])
    for membership in memberships:
        row: List[str | bool | None] = []
        for column, _ in COLUMN_DEFINITIONS:
            if column == 'email':
                row.append(membership.user.username)
            elif column == 'project':
                row.append(membership.project.name)
            elif column == 'title':
                row.append(membership.title or '')
            elif column in BOOLEAN_COLUMNS:
                row.append(bool(getattr(membership, column)))
            else:
                row.append('')
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _normalise_boolean(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().casefold()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return False


def _clean_email(value: object) -> str:
    email = str(value).strip().lower()
    if not email:
        raise ValidationError('Email is required')
    validate_email(email)
    return email


def import_memberships_workbook(
    workbook_file,
    *,
    accessible_projects: Sequence[Project],
) -> WorkbookImportResult:
    """Create or update memberships based on the uploaded workbook."""

    _require_openpyxl()
    result = WorkbookImportResult()
    try:
        wb = load_workbook(workbook_file, data_only=True)
    except Exception as exc:  # pragma: no cover - pass through
        raise MembershipWorkbookError('Unable to read the uploaded workbook.') from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise MembershipWorkbookError('The workbook is empty.')
    header = [str(value).strip() if value is not None else '' for value in rows[0]]
    expected_header = [label for _, label in COLUMN_DEFINITIONS]
    if header[: len(expected_header)] != expected_header:
        raise MembershipWorkbookError('The workbook headers do not match the expected template.')

    project_lookup: Mapping[str, Project] = {
        project.name.casefold(): project for project in accessible_projects
    }

    pending: dict[Tuple[str, int], dict[str, object]] = {}
    for row_index, row in enumerate(rows[1:], start=2):
        if row is None or not any(row):
            continue
        try:
            email = _clean_email(row[0])
        except ValidationError:
            result.errors.append(f'Row {row_index}: invalid email value.')
            continue
        project_raw = row[1] if len(row) > 1 else ''
        project_name = str(project_raw).strip()
        if not project_name:
            result.errors.append(f'Row {row_index}: project name is required.')
            continue
        project = project_lookup.get(project_name.casefold())
        if not project:
            result.errors.append(f'Row {row_index}: project "{project_name}" is not available.')
            continue
        payload: dict[str, object] = {
            'project': project,
            'title': (str(row[2]).strip() if len(row) > 2 and row[2] else ''),
        }
        for idx, (column, _) in enumerate(COLUMN_DEFINITIONS[3:], start=3):
            value = row[idx] if idx < len(row) else None
            payload[column] = _normalise_boolean(value)
        pending[(email, project.pk)] = payload

    if not pending:
        if result.errors:
            raise MembershipWorkbookError('No valid rows were found in the workbook.')
        raise MembershipWorkbookError('The workbook does not include any memberships.')

    with transaction.atomic():
        for (email, _project_pk), payload in pending.items():
            try:
                user = User.objects.get(username=email)
            except User.DoesNotExist:
                result.errors.append(f'User {email} does not exist and was skipped.')
                continue
            project = payload['project']  # type: ignore[assignment]
            membership_values = {key: value for key, value in payload.items() if key != 'project'}
            membership, created = Membership.objects.get_or_create(
                user=user,
                project=project,
                defaults=membership_values,
            )
            if created:
                result.created += 1
            else:
                for key, value in membership_values.items():
                    setattr(membership, key, value)
                membership.save()
                result.replaced += 1
            if membership.is_owner:
                (
                    Membership.objects.filter(project=project)
                    .exclude(pk=membership.pk)
                    .update(is_owner=False)
                )
    return result
