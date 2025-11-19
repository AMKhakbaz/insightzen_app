"""Helpers for ingesting Excel-based respondent samples per project."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Sequence, Tuple

from django.db import transaction
from django.utils import timezone

try:  # pragma: no cover - optional dependency guarded at runtime
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - openpyxl always available in prod, but guard for tests
    load_workbook = None

from ..models import CallSample, Mobile, Person, Project, UploadedSampleEntry
from .gender_utils import normalize_gender_value

SAMPLE_REQUIRED_HEADERS = ['full_name', 'phone', 'city', 'age', 'gender']
HEADER_ALIASES = {
    'name': 'full_name',
    'full name': 'full_name',
    'fullname': 'full_name',
    'mobile': 'phone',
    'mobile_number': 'phone',
    'city_name': 'city',
    'province': 'city',
    'age_years': 'age',
    'sex': 'gender',
}


class SampleUploadError(Exception):
    """Raised when an uploaded workbook cannot be processed."""


@dataclass
class SampleUploadStats:
    total_rows: int
    accepted_rows: int

    @property
    def skipped_rows(self) -> int:
        return max(self.total_rows - self.accepted_rows, 0)


@dataclass
class SampleAppendResult:
    stats: SampleUploadStats
    appended_rows: int
    duplicate_rows: int = 0
    rejected_rows: int = 0


@dataclass
class ParsedSampleRow:
    full_name: str
    phone: str
    city: str
    age: int | None
    gender: str
    created_from_row: int


def _normalise_header(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ''


def _normalise_phone(value: Any) -> str:
    text = str(value).strip() if value is not None else ''
    # Remove trailing .0 that Excel might append to numeric strings
    if text.endswith('.0') and text.replace('.', '', 1).isdigit():
        text = text[:-2]
    return text


def _coerce_int(value: Any) -> int | None:
    if value in (None, ''):
        return None
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _coerce_text(value: Any) -> str:
    return str(value).strip() if value not in (None, '') else ''


def _load_workbook(project: Project, uploaded_file=None):
    if load_workbook is None:
        raise SampleUploadError('Excel support is not available on this server.')
    upload = uploaded_file or project.sample_upload
    if not upload:
        raise SampleUploadError('No workbook is attached to this project.')
    should_close = False
    if uploaded_file is None:
        upload.open('rb')
        should_close = True
    else:
        try:
            upload.seek(0)
        except Exception:
            pass
    try:
        try:
            workbook = load_workbook(upload, read_only=True, data_only=True)
        except Exception as exc:  # pragma: no cover
            raise SampleUploadError('The uploaded workbook could not be read.') from exc
        return workbook, should_close, upload
    except Exception:
        if should_close:
            upload.close()
        raise


def _parse_uploaded_sample_rows(project: Project, uploaded_file=None) -> Tuple[List[ParsedSampleRow], SampleUploadStats]:
    workbook, should_close, upload_handle = _load_workbook(project, uploaded_file)
    try:
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise SampleUploadError('The uploaded workbook does not contain any rows.')

        header_map: Dict[str, int] = {}
        for idx, raw_header in enumerate(header_row):
            header_name = HEADER_ALIASES.get(_normalise_header(raw_header), _normalise_header(raw_header))
            if header_name:
                header_map[header_name] = idx

        missing = [col for col in SAMPLE_REQUIRED_HEADERS if col not in header_map]
        if missing:
            raise SampleUploadError('Missing required columns: ' + ', '.join(missing))

        seen_numbers: set[str] = set()
        rows: List[ParsedSampleRow] = []
        total_rows = 0
        for excel_row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            total_rows += 1
            phone = _normalise_phone(row[header_map['phone']] if header_map.get('phone') is not None else None)
            if not phone or phone in seen_numbers:
                continue
            seen_numbers.add(phone)
            parsed_row = ParsedSampleRow(
                phone=phone,
                full_name=_coerce_text(row[header_map['full_name']]) if header_map.get('full_name') is not None else '',
                city=_coerce_text(row[header_map['city']]) if header_map.get('city') is not None else '',
                age=_coerce_int(row[header_map['age']]) if header_map.get('age') is not None else None,
                gender=normalize_gender_value(
                    row[header_map['gender']] if header_map.get('gender') is not None else None
                ) or '',
                created_from_row=excel_row_index,
            )
            rows.append(parsed_row)

        if not rows:
            raise SampleUploadError('No valid rows were found in the uploaded workbook.')

        stats = SampleUploadStats(total_rows=total_rows, accepted_rows=len(rows))
        return rows, stats
    finally:
        if should_close:
            upload_handle.close()


def _build_uploaded_entries(project: Project, rows: Sequence[ParsedSampleRow]) -> List[UploadedSampleEntry]:
    entries: List[UploadedSampleEntry] = []
    for row in rows:
        entries.append(
            UploadedSampleEntry(
                project=project,
                phone=row.phone,
                full_name=row.full_name,
                city=row.city,
                age=row.age,
                gender=row.gender,
                metadata={'source': 'upload'},
                created_from_row=row.created_from_row,
            )
        )
    return entries


@transaction.atomic
def ingest_project_sample_upload(project: Project) -> SampleUploadStats:
    """Parse the project's uploaded workbook and materialise sample entries."""

    rows, stats = _parse_uploaded_sample_rows(project)
    entries = _build_uploaded_entries(project, rows)
    UploadedSampleEntry.objects.filter(project=project).delete()
    UploadedSampleEntry.objects.bulk_create(entries, batch_size=500)

    project.sample_upload_refreshed_at = timezone.now()
    project.sample_upload_metadata = {
        'total_rows': stats.total_rows,
        'accepted_rows': stats.accepted_rows,
        'skipped_rows': stats.skipped_rows,
        'appended_rows': stats.accepted_rows,
        'duplicate_rows': stats.skipped_rows,
        'required_headers': SAMPLE_REQUIRED_HEADERS,
    }
    project.save(update_fields=['sample_upload_refreshed_at', 'sample_upload_metadata'])
    return stats


def _collect_existing_numbers(project: Project) -> set[str]:
    existing_numbers = set(
        UploadedSampleEntry.objects.filter(project=project).values_list('phone', flat=True)
    )
    call_sample_numbers = set(
        CallSample.objects.filter(project=project, mobile__isnull=False).values_list('mobile__mobile', flat=True)
    )
    existing_numbers |= call_sample_numbers
    return existing_numbers


@transaction.atomic
def append_project_sample_upload(project: Project, uploaded_file=None) -> SampleAppendResult:
    rows, stats = _parse_uploaded_sample_rows(project, uploaded_file=uploaded_file)
    existing_numbers = _collect_existing_numbers(project)
    unique_rows: List[ParsedSampleRow] = []
    duplicate_rows = 0
    for row in rows:
        if row.phone in existing_numbers:
            duplicate_rows += 1
            continue
        unique_rows.append(row)
        existing_numbers.add(row.phone)
    if not unique_rows:
        raise SampleUploadError('All rows were skipped because the phone numbers already exist.')
    entries = _build_uploaded_entries(project, unique_rows)
    UploadedSampleEntry.objects.bulk_create(entries, batch_size=500, ignore_conflicts=True)
    appended = len(entries)
    project.sample_upload_refreshed_at = timezone.now()
    now = timezone.now()
    project.sample_upload_refreshed_at = now
    project.sample_upload_metadata = {
        'total_rows': stats.total_rows,
        'accepted_rows': stats.accepted_rows,
        'skipped_rows': stats.skipped_rows,
        'appended_rows': appended,
        'duplicate_rows': duplicate_rows,
        'required_headers': SAMPLE_REQUIRED_HEADERS,
    }
    project.save(update_fields=['sample_upload_refreshed_at', 'sample_upload_metadata'])
    return SampleAppendResult(stats=stats, appended_rows=appended, duplicate_rows=duplicate_rows)


def _generate_person_code(project_id: int, phone: str) -> str:
    digits = ''.join(ch for ch in phone if ch.isdigit()) or '0'
    suffix = digits[-6:].rjust(6, '0')
    prefix = str(project_id % 1000).rjust(3, '0')
    return f'P{prefix}{suffix}'[:10]


@transaction.atomic
def append_project_respondent_bank(project: Project, uploaded_file=None) -> SampleAppendResult:
    rows, stats = _parse_uploaded_sample_rows(project, uploaded_file=uploaded_file)
    phones_to_check = {row.phone for row in rows}
    existing_mobile_numbers = set(
        Mobile.objects.filter(mobile__in=phones_to_check).values_list('mobile', flat=True)
    )
    existing_numbers = existing_mobile_numbers | _collect_existing_numbers(project)
    new_persons: List[Person] = []
    new_mobiles: List[Mobile] = []
    appended_rows = 0
    duplicate_rows = 0
    base_year = timezone.now().year
    for row in rows:
        if row.phone in existing_numbers:
            duplicate_rows += 1
            continue
        age_value = row.age if row.age is not None else 30
        birth_year = base_year - age_value
        national_code = _generate_person_code(project.pk, row.phone)
        new_persons.append(
            Person(
                national_code=national_code,
                full_name=row.full_name or row.phone,
                birth_year=birth_year,
                city_name=row.city or 'Unspecified',
                gender=row.gender or None,
            )
        )
        new_mobiles.append(
            Mobile(
                mobile=row.phone,
                person_id=national_code,
            )
        )
        appended_rows += 1
        existing_numbers.add(row.phone)

    if not appended_rows:
        raise SampleUploadError('All phone numbers already exist in the respondent bank.')

    Person.objects.bulk_create(new_persons, ignore_conflicts=True, batch_size=500)
    Mobile.objects.bulk_create(new_mobiles, ignore_conflicts=True, batch_size=500)
    now = timezone.now()
    project.sample_upload_refreshed_at = now
    project.sample_upload_metadata = {
        'total_rows': stats.total_rows,
        'accepted_rows': stats.accepted_rows,
        'skipped_rows': stats.skipped_rows,
        'appended_rows': appended_rows,
        'duplicate_rows': duplicate_rows,
        'required_headers': SAMPLE_REQUIRED_HEADERS,
    }
    project.save(update_fields=['sample_upload_metadata', 'sample_upload_refreshed_at'])
    return SampleAppendResult(stats=stats, appended_rows=appended_rows, duplicate_rows=duplicate_rows)
