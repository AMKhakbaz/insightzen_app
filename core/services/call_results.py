"""Utilities for managing call result code definitions and uploads."""

from __future__ import annotations

import csv
from collections import OrderedDict
from dataclasses import dataclass
from io import TextIOWrapper
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from django.db import transaction
from django.utils import timezone

try:  # pragma: no cover - optional dependency guarded for environments without Excel libs
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None

from ..models import Project, ProjectCallResult

CALL_RESULT_REQUIRED_HEADERS = ['code', 'label', 'is_success']
HEADER_ALIASES = {
    'codes': 'code',
    'result_code': 'code',
    'status_code': 'code',
    'title': 'label',
    'name': 'label',
    'description': 'label',
    'success': 'is_success',
    'is success': 'is_success',
    'successful': 'is_success',
}
TRUE_VALUES = {'1', 'true', 't', 'yes', 'y'}
FALSE_VALUES = {'0', 'false', 'f', 'no', 'n'}


class CallResultUploadError(Exception):
    """Raised when a call result upload cannot be processed."""


@dataclass
class CallResultUploadStats:
    total_rows: int
    accepted_rows: int

    @property
    def skipped_rows(self) -> int:
        return max(self.total_rows - self.accepted_rows, 0)


@dataclass(frozen=True)
class DefaultCallResultDefinition:
    code: int
    label_en: str
    label_fa: str
    is_success: bool = False


DEFAULT_CALL_RESULTS: List[DefaultCallResultDefinition] = [
    DefaultCallResultDefinition(1, 'Successful Interview', 'مصاحبه موفق', True),
    DefaultCallResultDefinition(2, 'Voicemail', 'پیغام گیر (صندوق صوتی)'),
    DefaultCallResultDefinition(3, 'Call later (with time)', 'بعدا تماس بگیرید (با تعیین زمان)'),
    DefaultCallResultDefinition(4, 'Call later', 'بعدا تماس بگیرید (بدون تعیین زمان)'),
    DefaultCallResultDefinition(5, 'Busy', 'اشغال است'),
    DefaultCallResultDefinition(6, 'No answer', 'جواب نمی‌دهد'),
    DefaultCallResultDefinition(7, 'Incomplete interview (to be completed)', 'مصاحبه ناقص (باید تکمیل شود)'),
    DefaultCallResultDefinition(8, 'Incomplete (respondent unwilling)', 'مصاحبه ناقص (تمایلی به ادامه ندرد)'),
    DefaultCallResultDefinition(9, 'Number not in network', 'شماره در شبکه موجود نیست'),
    DefaultCallResultDefinition(10, 'Language barrier', 'مشکل زبان'),
    DefaultCallResultDefinition(11, 'Respondent unavailable during fieldwork', 'پاسخگو در مدت فیلد در دسترس نیست'),
    DefaultCallResultDefinition(12, 'Powered off', 'خاموش است'),
    DefaultCallResultDefinition(13, 'Non‑cooperative', 'عدم همکاری'),
    DefaultCallResultDefinition(14, 'Do not call again (angry)', 'دیگر تماس نگیرید (پاسخگوی عصبانی)'),
    DefaultCallResultDefinition(15, 'Not eligible', 'پاسخگوی غیر واجد شرایط'),
    DefaultCallResultDefinition(16, 'Quota exceeded', 'بیش از سهمیه'),
    DefaultCallResultDefinition(17, 'Unavailable', 'در دسترس نیست'),
    DefaultCallResultDefinition(18, 'Cannot connect', 'برقراری تماس مقدور نیست'),
    DefaultCallResultDefinition(19, 'Out of service', 'خارج از سرویس'),
    DefaultCallResultDefinition(20, 'Other', 'سایر'),
    DefaultCallResultDefinition(21, 'Burned interview', 'مصاحبه سوخته'),
]


@dataclass
class CallResultDefinitions:
    labels: "OrderedDict[int, str]"
    success_map: Dict[int, bool]
    source: str  # 'default' or 'custom'


def _normalise_header(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ''


def _coerce_int(value: Any) -> Optional[int]:
    if value in (None, ''):
        return None
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _coerce_label(value: Any) -> str:
    return str(value).strip() if value not in (None, '') else ''


def _coerce_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value in (None, ''):
        return False
    text = str(value).strip().lower()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    # default to False but raise explicit error elsewhere when validation fails
    raise ValueError


def _build_error(message_en: str, message_fa: str) -> str:
    return f"{message_en} / {message_fa}"


@transaction.atomic
def ingest_project_call_result_upload(project: Project) -> CallResultUploadStats:
    """Parse the project's uploaded CSV/Excel file into call result rows."""

    upload = project.call_result_upload
    if not upload:
        raise CallResultUploadError(_build_error('No file has been uploaded for this project.', 'فایلی برای این پروژه بارگذاری نشده است.'))

    suffix = Path(upload.name or '').suffix.lower()
    rows: List[Iterable[Any]] = []
    header_row: List[Any] = []

    if suffix in {'.xlsx', '.xls'}:
        if load_workbook is None:
            raise CallResultUploadError(_build_error('Excel support is not available on this server.', 'امکان پردازش فایل اکسل روی این سرور فعال نیست.'))
        upload.open('rb')
        try:
            workbook = load_workbook(upload, read_only=True, data_only=True)
        except Exception as exc:  # pragma: no cover - delegated to UI feedback
            raise CallResultUploadError(_build_error('The uploaded workbook could not be read.', 'فایل اکسل بارگذاری‌شده قابل خواندن نیست.')) from exc
        finally:
            upload.close()
        worksheet = workbook.active
        header_row = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), []))
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            rows.append(list(row))
    elif suffix == '.csv':
        upload.open('rb')
        try:
            wrapper = TextIOWrapper(upload, encoding='utf-8-sig')
            reader = list(csv.reader(wrapper))
        except Exception as exc:  # pragma: no cover
            raise CallResultUploadError(_build_error('The uploaded CSV file could not be read.', 'امکان خواندن فایل CSV بارگذاری‌شده وجود ندارد.')) from exc
        finally:
            upload.close()
        if reader:
            header_row = reader[0]
            rows = [row for row in reader[1:] if row]
    else:
        raise CallResultUploadError(_build_error('Unsupported file format. Please upload CSV or Excel files.', 'فرمت فایل پشتیبانی نمی‌شود. لطفاً CSV یا اکسل بارگذاری کنید.'))

    if not header_row:
        raise CallResultUploadError(_build_error('The uploaded file does not contain a header row.', 'فایل بارگذاری‌شده فاقد ردیف سرستون است.'))

    header_map: Dict[str, int] = {}
    for idx, raw_header in enumerate(header_row):
        header_name = HEADER_ALIASES.get(_normalise_header(raw_header), _normalise_header(raw_header))
        if header_name:
            header_map[header_name] = idx

    missing = [col for col in CALL_RESULT_REQUIRED_HEADERS if col not in header_map]
    if missing:
        raise CallResultUploadError(
            _build_error(
                'Missing required columns: ' + ', '.join(missing),
                'ستون‌های ضروری یافت نشد: ' + ', '.join(missing),
            )
        )

    seen_codes: set[int] = set()
    entries: List[ProjectCallResult] = []
    total_rows = 0
    for row in rows:
        total_rows += 1
        code_idx = header_map['code']
        label_idx = header_map['label']
        success_idx = header_map['is_success']
        code_val = _coerce_int(row[code_idx] if len(row) > code_idx else None)
        if code_val is None or code_val in seen_codes:
            continue
        label_val = _coerce_label(row[label_idx] if len(row) > label_idx else None)
        if not label_val:
            continue
        success_raw = row[success_idx] if len(row) > success_idx else None
        try:
            success_val = _coerce_bool(success_raw)
        except ValueError:
            raise CallResultUploadError(
                _build_error(
                    'The "is_success" column must contain yes/no or 0/1 values.',
                    'ستون is_success باید شامل مقادیر بلی/خیر یا ۰/۱ باشد.',
                )
            )
        seen_codes.add(code_val)
        entries.append(
            ProjectCallResult(
                project=project,
                code=code_val,
                label=label_val,
                is_success=success_val,
                display_order=len(entries),
            )
        )

    if not entries:
        raise CallResultUploadError(_build_error('No valid rows were found in the uploaded file.', 'هیچ ردیف معتبری در فایل بارگذاری‌شده یافت نشد.'))

    ProjectCallResult.objects.filter(project=project).delete()
    ProjectCallResult.objects.bulk_create(entries, batch_size=200)

    stats = CallResultUploadStats(total_rows=total_rows, accepted_rows=len(entries))
    project.call_result_metadata = {
        'total_rows': stats.total_rows,
        'accepted_rows': stats.accepted_rows,
        'skipped_rows': stats.skipped_rows,
        'required_headers': CALL_RESULT_REQUIRED_HEADERS,
    }
    project.call_result_refreshed_at = timezone.now()
    project.save(update_fields=['call_result_metadata', 'call_result_refreshed_at'])
    return stats


def clear_project_call_results(project: Project) -> None:
    """Remove any stored custom call result definitions and reset metadata."""

    ProjectCallResult.objects.filter(project=project).delete()
    if project.call_result_upload:
        project.call_result_upload.delete(save=False)
    project.call_result_upload = None
    project.call_result_metadata = {}
    project.call_result_refreshed_at = None
    project.save(update_fields=['call_result_upload', 'call_result_metadata', 'call_result_refreshed_at'])


def resolve_call_result_definitions(project: Optional[Project], lang: str = 'en') -> CallResultDefinitions:
    """Return ordered labels and success map for the given project."""

    lang = 'fa' if lang == 'fa' else 'en'
    if project and project.call_result_source == Project.CallResultSource.CUSTOM:
        rows = list(project.call_results.order_by('display_order', 'code'))
        if rows:
            labels = OrderedDict((row.code, row.label) for row in rows)
            success_map = {row.code: row.is_success for row in rows}
            return CallResultDefinitions(labels=labels, success_map=success_map, source='custom')

    labels = OrderedDict()
    success_map: Dict[int, bool] = {}
    for definition in DEFAULT_CALL_RESULTS:
        label = definition.label_fa if lang == 'fa' else definition.label_en
        labels[definition.code] = label
        success_map[definition.code] = definition.is_success
    return CallResultDefinitions(labels=labels, success_map=success_map, source='default')


def code_is_successful(project: Optional[Project], code: Optional[int], definitions: Optional[CallResultDefinitions] = None) -> bool:
    """Return True if the provided code represents a successful interview."""

    if code is None:
        return False
    if definitions is None:
        definitions = resolve_call_result_definitions(project)
    return definitions.success_map.get(code, False)
